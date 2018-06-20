# -*- coding: utf-8 -*-
"""
Created on Fri Jun  1 12:02:45 2018

@author: Joshua Hibbard

"""

import re
import csv
import os
import shutil
import collections
import numpy as np
from openpyxl import *
from openpyxl.utils import coordinate_from_string, column_index_from_string
from openpyxl.workbook import Workbook


date = input('What is the date of the experiment? Uses spaces to separate the year, month, and date. Format: YYYY MM DD') 

start_time = input('What is the time stamp of the first image? Please use the format HH MM SS MSS and separate each entry with a space.')

R = input('What is the R value? (Radiation load)')


date = date.split()
year=date[0]
month=date[1]
day=date[2]

Main_Directory = 'C:\\Users\\Undergrunt\\Box\\Josh data\\Red_Blue Light\\' + month + day + year[-2:]
os.chdir(Main_Directory)


start_time_list = start_time.split()
hours=start_time_list[0]
minutes=start_time_list[1]
seconds=start_time_list[2]
milliseconds=start_time_list[3]
thermal_image_time_list=[]
"Delineates the columns in the gas exchange data with the desired values."
time_stamp_column=2
lower_before_thermo_column=26
lower_after_thermo_column=27
upper_before_thermo_column=28
upper_after_thermo_column=29
xout2_column=22
"Convert the time strings into numbers."
image_true_hours=float(hours)
image_true_minutes=float(minutes)
image_true_seconds=float(seconds)
image_true_milliseconds=float(milliseconds)
'Convert the true image time into minutes elapsed since midnight.'
first_image_time = image_true_hours*60 + image_true_minutes + image_true_seconds/60 + image_true_milliseconds/60000


list1 = os.listdir('OriginalTempImages')
number_files = len(list1)

minutes_between_images = 3
image_index = 0
"Create an array of all the image times."
while image_index < number_files:
    final_image_time = minutes_between_images*image_index + first_image_time
    thermal_image_time_list.append(final_image_time)
    image_index+=1
    

thermal_image_index = 0
file_number_index = 0
list_index = 1
d={}
data_extraction = {}

"Find the time stamps of each image and place them into a dictionary."
for filename in os.listdir('OriginalTempImages'):
    regex=re.compile('([0-9]*)(?:\\.csv)')
    time_stamp1=regex.findall(filename)
    time_stamp1=time_stamp1[0]
    time_stamp1=float(time_stamp1)
    d[time_stamp1]=filename
    
'Create the folders for output'
os.makedirs(Main_Directory + '\\' + 'FinalTempImages')
os.makedirs(Main_Directory + '\\' + 'TimeStampedTempImages')
os.makedirs(Main_Directory + '\\' + 'Conductance Maps')

od = collections.OrderedDict(sorted(d.items()))

time_stamp_list = []

with open(year + '_' + month + '_' + day + '.csv', 'r') as gas_exchange_data, open('DataExtraction.csv','w') as outputfile:
    data = csv.reader(gas_exchange_data, delimiter = ',', quotechar = '\n')
    
    while thermal_image_index < number_files:
        get_row = next(data)
        time_stamp=float(get_row[time_stamp_column])
        if time_stamp <= thermal_image_time_list[thermal_image_index] + 0.3 and time_stamp >= thermal_image_time_list[thermal_image_index] - 0.3:
            
            lbt=float(get_row[lower_before_thermo_column])
            lat=float(get_row[lower_after_thermo_column])
            ubt=float(get_row[upper_before_thermo_column])
            uat=float(get_row[upper_after_thermo_column])
            xout2=float(get_row[xout2_column])
            
            data_extraction[time_stamp] = [ubt, uat, lbt, lat, xout2]
        
            outputfile.write(str(time_stamp))
            outputfile.write(',')
            outputfile.write('KMatrix_23C_oneamp')
            outputfile.write(',')
            outputfile.write(str(ubt))
            outputfile.write(',')
            outputfile.write(str(uat))
            outputfile.write(',')
            outputfile.write(str(lbt))
            outputfile.write(',')
            outputfile.write(str(lat))
            outputfile.write(',')
            outputfile.write(str(xout2))
            outputfile.write('\n')
        
            thermal_image_index+=1        
            
                
            if file_number_index < number_files:
                filename = od[list_index]
                shutil.copy(Main_Directory + '\\OriginalTempImages\\' + filename, Main_Directory + '\\TimeStampedTempImages\\'+str(time_stamp)+'.csv')
                time_stamp_list.append(time_stamp)
                file_number_index+=1
                list_index+=1



os.chdir(Main_Directory + '\\TimeStampedTempImages')
'Crops the images to include only the leaf.'
for filename in os.listdir(Main_Directory + '\\TimeStampedTempImages'):
    

    with open(filename,'r', encoding='utf-8-sig') as T_ecsvfile, open('Leaf Temperature ' + filename, 'w', newline = '') as LTFinal:
        in_file = csv.reader(T_ecsvfile)
        out_file = csv.writer(LTFinal)
        for row_number, row in enumerate(in_file):
            if row_number > 80 and row_number < 414:
                
                out_file.writerow(row[174:568])
    shutil.copy(Main_Directory + '\\TimeStampedTempImages\\' + 'Leaf Temperature ' + filename, Main_Directory + '\\FinalTempImages')

os.chdir(Main_Directory)



'Create Excel workbook to output the pixel data to.'
workbook = load_workbook('Conductance Graphs.xlsx')
ws = workbook.create_sheet(''.join(date))
ws.cell(row = 1, column = 1).value = 'Red/Blue Light Experiments'
ws.cell(row = 3, column = 1).value = 'Date'
ws.cell(row = 4, column = 1).value = ''.join(date)
ws.cell(row = 1, column = 3).value = 'Air Temperature'
ws.cell(row = 1, column = 4).value = '23 C'







pixel_list = input('What are the Excel Coordinates you would like to analyze? Separate each coordinate with a space.')
pixel_list = pixel_list.split()




"Constants"

"Latent Heat of Vaporization for Water"
L_w = 40.68
w_0 = 657959000
T_w = 4982.85
R=float(R)

"Calculate Conductance."
def g_s(K_matrix,T_a,T_e):
    return (R + K_matrix*(T_a-T_e))/(L_w*(w_0*np.exp(-T_w/(T_e+273)) - w_a))
"Calculate Air Temperature."
def T_a(columnindex):
    return ((after_average-before_average)/(rowlength-1))*columnindex + before_average
'Calculate Delta W'
def delta_w(T_e_at_coordinate):
    return w_0*np.exp(-T_w/(T_e+273)) - w_a



for filename in os.listdir('FinalTempImages'):
    "Find air temperatures from the data_extraction file."
    ubt = data_extraction[filename[-17:-4]][0]
    uat = data_extraction[filename[-17:-4]][1]
    lbt = data_extraction[filename[-17:-4]][2]
    lat = data_extraction[filename[-17:-4]][3]
    "Calculate average before and average after temperatures from the thermocouples."
    before_average = (ubt + lbt)/2
    after_average = (uat+lat)/2
    "Find water mole fraction from data extraction."
    w_a = data_extraction[filename[-17:-4]][4]
    print(w_a)
    with open('C:\\Users\\Joshua Hibbard\\Box\\Josh data\\KMatrix\\KMatrix_23C_1Amp.csv','r') as K_matrixcsvfile, open(filename,'r', encoding='utf-8-sig') as T_ecsvfile, open(year + '_' + month +'_' + day + '_' + 'Conductance' + filename[-17:] + '.csv','w') as outputfile:
        K_matrix = csv.reader(K_matrixcsvfile, delimiter = ',', quotechar='\n')
        T_e = csv.reader(T_ecsvfile, delimiter = ',', quotechar='\n')
        while True:
            K_matrixrow=next(K_matrix)
            T_erow=next(T_e)
            rowlength=len(T_erow)
            columnindex=0
            while columnindex < rowlength:
                K_matrixvalue = float(K_matrixrow[columnindex])
                T_avalue = float(T_a(columnindex))
                T_evalue = float(T_erow[columnindex])
                conductance = g_s(K_matrixvalue,T_avalue,T_evalue)
                outputfile.write(str(conductance))
                columnindex+=1
                if columnindex < rowlength:
                    outputfile.write(',')
            outputfile.write('\n')
    shutil.copy(Main_Directory + '\\FinalTempImages\\' + year + '_' + month +'_' + day + '_' + 'Conductance_' + filename[-17:] + '.csv', Main_Directory + '\\Conductance Maps\\')


Titles = 'yes'
k = 0
l = 0


for filename in os.listdir('TimeStampedTempImages'):
     with open(Main_Directory + '\\Conductance Maps\\' + year + '_' + month +'_' + day + '_' + 'Conductance_' + filename,'r') as Conductance, open(Main_Directory + '\\FinalTempImages\\' + 'Leaf Temperature ' + filename,'r', encoding='utf-8-sig') as T_ecsvfile:
        conductance_rows = list(csv.reader(Conductance))
        T_e_rows = list(csv.reader(T_ecsvfile))
        if Titles == 'yes':
            for pixel in pixel_list():
                ws.cell(row = 6 + k*number_files + l, column = 1).value = 'Excel Coordinate'
                ws.cell(row = 6 + k*number_files + l, column = 2).value = pixel
                ws.cell(row = 6 + k*number_files + l, column = 1).value = 'Image Time'
                ws.cell(row = 6 + k*number_files + l, column = 2).value = 'w_a/Xout2'
                ws.cell(row = 6 + k*number_files + l, column = 4).value = 'Pixel Temperature'
                ws.cell(row = 6 + k*number_files + l, column = 5).value = 'Pixel Delta W'
                ws.cell(row = 6 + k*number_files + l, column = 6).value = 'Pixel Conductance'
                ws.cell(row = 6 + k*number_files + l, column = 8).value = 'Leaflet Temperature'
                ws.cell(row = 6 + k*number_files + l, column = 9).value = 'Leaflet Delta W'
                ws.cell(row = 6 + k*number_files + l, column = 10).value = 'Leaflet Conductance'
                l+=1
                k+=2
        Titles = 'no'
        k = 0
        l = 0
        for pixel in pixel_list():
            xy = coordinate_from_string(pixel)
            col = column_index_from_string(xy[0])
            row = xy[1]
            conductance_at_coordinate = conductance[row][col]
            T_e_at_coordinate = T_e_rows[row][col]
            aa = T_e_rows[row - 1][col - 1]
            ab = T_e_rows[row - 1][col]
            ac = T_e_rows[row - 1][col + 1]
            ba = T_e_rows[row][col - 1]
            bc = T_e_rows[row][col + 1]
            ca = T_e_rows[row + 1][col - 1]
            cb = T_e_rows[row + 1][col]
            cc = T_e_rows[row + 1][col + 1]
            leaflet_temperature = (aa + ab + ac + ba + T_e_at_coordinate + bc + ca + cb +cc)/9
            
            aa1 = conductance[row - 1][col - 1]
            ab1 = conductance[row - 1][col]
            ac1 = conductance[row - 1][col + 1]
            ba1 = conductance[row][col - 1]
            bc1 = conductance[row][col + 1]
            ca1 = conductance[row + 1][col - 1]
            cb1 = conductance[row + 1][col]
            cc1 = conductance[row + 1][col + 1]
            leaflet_conductance = (aa1 + ab1 + ac1 + ba1 + conductance_at_coordinate + bc1 + ca1 + cb1+ cc1)/9
            
            ws.cell(row = 6 + k*number_files + l, column = 1).value = filename[:-4]
            ws.cell(row = 6 + k*number_files + l, column = 2).value = data_extraction[filename[:-4]][4]
            ws.cell(row = 6 + k*number_files + l, column = 4).value = T_e_at_coordinate
            ws.cell(row = 6 + k*number_files + l, column = 5).value = delta_w(T_e_at_coordinate)
            ws.cell(row = 6 + k*number_files + l, column = 6).value = conductance_at_coordinate
            ws.cell(row = 6 + k*number_files + l, column = 8).value = leaflet_temperature
            ws.cell(row = 6 + k*number_files + l, column = 9).value = delta_w(leaflet_temperature)
            ws.cell(row = 6 + k*number_files + l, column = 10).value = leaflet_conductance
            l+=1
            k+=2

workbook.save(filename = Main_Directory + 'Graph Analysis.xlsx')    
                
            
            
            
            
            
